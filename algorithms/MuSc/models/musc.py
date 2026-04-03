"""
translated by codegeex
1. 代码概述
MuSc 类是一个完整的异常检测系统，它使用预训练的视觉模型（如CLIP或DINO）作为特征提取器，
通过LNAMD（局部非负矩阵分解）、MSM（多尺度互评分）和RsCIN（基于残差的类别内归一化）等模块进行异常检测和评分。

2. 主要组件和功能
    2.1 初始化 (__init__)
        设置基本参数：设备（CPU/GPU）、数据集路径、类别、模型名称等
        支持三种数据集：VisA、MVTec AD和BTAD
        创建输出目录用于保存结果
    2.2 骨干网络加载 (load_backbone)
        支持两种预训练模型：
        DINO/DINOv2：用于提取视觉特征
        CLIP：多模态模型，用于图像编码
        将模型加载到指定设备（GPU/CPU）
    2.3 数据集加载 (load_datasets)
        根据配置加载指定的数据集类别
        支持数据集划分和迭代
        应用预处理转换器
    2.4 可视化 (visualization)
        两种可视化模式：
            单图归一化：每张图单独归一化
            全局归一化：所有图一起归一化
            使用JET色图生成热力图
    2.5 类别数据处理 (evaluate_category_data)
        这是核心处理流程，包括：
        特征提取：
            使用骨干网络提取图像特征和补丁特征
            支持多层级特征提取
        LNAMD处理：
            对提取的特征进行局部非负矩阵分解
            使用不同聚合度（r_list）处理特征
        MSM处理：
            多尺度互评分模块
            计算不同层级特征的异常分数
        特征插值：
            将异常图插值到原始图像大小
        RsCIN处理：
            基于残差的类别内归一化
            结合图像级特征进行分类
        指标计算：
            计算图像级和像素级的AUROC、F1、AP等指标
    2.6 主流程 (evaluate)
        遍历所有类别进行异常检测
        计算并保存各项指标
        将结果保存到Excel文件
3. 实现原理
    特征提取：
        使用预训练的视觉模型（CLIP或DINO）提取图像的多层级特征
        包括图像级特征和补丁级特征
    异常检测流程：
        LNAMD：对补丁特征进行局部非负矩阵分解，聚合局部特征
        MSM：通过多尺度互评分计算异常分数，捕捉不同尺度的异常
        RsCIN：结合图像级特征进行类别内归一化，提高分类准确性
    评分机制：
        图像级评分：基于异常图的最大值
        像素级评分：基于插值后的异常图
"""

import os
import time

import cv2
import numpy as np
import torch
import torch.nn.functional as F
from openpyxl import Workbook
from tqdm import tqdm

from MuSc.dataset import mvtec
from MuSc.dataset.mvtec import CLASSNAMES
from MuSc.models.backbone import _backbones
from MuSc.models.backbone import open_clip
from MuSc.models.modules.LNAMD import LNAMD
from MuSc.models.modules.MSM import MSM
from MuSc.models.modules.RsCIN import RsCIN
from MuSc.utils.metrics import compute_metrics


class MuSc:
    def __init__(self, cfg, seed=0):
        self.cfg = cfg
        self.seed = seed
        # 支持CPU模式：如果cfg['device']是'cpu'字符串，则使用CPU
        if isinstance(cfg['device'], str) and cfg['device'].lower() == 'cpu':
            self.device = torch.device("cpu")
        elif torch.cuda.is_available():
            self.device = torch.device("cuda:{}".format(cfg['device']))
        else:
            self.device = torch.device("cpu")
        self.path = cfg['datasets']['data_path']
        self.dataset = cfg['datasets']['dataset_name']
        self.vis = cfg['testing']['vis']
        self.vis_type = cfg['testing']['vis_type']
        self.save_excel = cfg['testing']['save_excel']
        # the categories to be tested
        self.categories = cfg['datasets']['class_name']
        if isinstance(self.categories, str):
            if self.categories.lower() == 'all':
                if self.dataset == 'mvtec_ad':
                    self.categories = CLASSNAMES
            else:
                self.categories = [self.categories]
        # model setting
        self.model_name = cfg['models']['backbone_name']
        self.pretrained_models_dir = cfg['models']['pretrained_models_dir']
        self.dino_model = None
        self.preprocess = None
        self.clip_model = None
        self.pretrained = cfg['models'][self.model_name]['pretrained']
        # params
        self.image_size = cfg['models'][self.model_name]['img_resize']
        self.batch_size = cfg['models']['batch_size']
        self.r_list = cfg['models']['r_list']
        self.features_list = [l + 1 for l in cfg['models'][self.model_name]['feature_layers']]
        self.divide_num = cfg['datasets']['divide_num']
        self.output_dir = str(os.path.join(cfg['testing']['output_dir'], self.dataset, self.model_name,
                                           'imagesize{}'.format(self.image_size)))
        os.makedirs(self.output_dir, exist_ok=True)
        self.load_backbone()

    def load_backbone(self):
        """
        加载预训练的骨干网络
        """
        if 'dino' in self.model_name:
            # dino or dino_v2
            self.dino_model = _backbones.load(self.model_name, model_cache_dir=self.pretrained_models_dir)
            self.dino_model.to(self.device)
            self.preprocess = None
        else:
            # clip
            self.clip_model, _, self.preprocess = open_clip.create_model_and_transforms(self.model_name,
                                                                                        self.image_size,
                                                                                        pretrained=self.pretrained,
                                                                                        cache_dir=self.pretrained_models_dir)
            self.clip_model.to(self.device)

    def load_datasets(self, category, divide_num=1, divide_iter=0):
        """
        加载指定类别的数据集
        参数:
            category (str): 要加载的数据集类别名称
            divide_num (int, optional): 数据集划分的数量，默认为1
            divide_iter (int, optional): 数据集划分的迭代次数，默认为0
        返回:
            根据self.dataset指定的类型返回相应的测试数据集对象
        注意:
            目前支持三种数据集类型：'visa'、'mvtec_ad'和'btad'
            每种数据集都会使用相同的参数进行初始化，包括数据源、分割方式、类别名称、图像大小、预处理等
        """
        # dataloader
        test_dataset = None
        if self.dataset == 'mvtec_ad':
            test_dataset = mvtec.MVTecDataset(source=self.path, split=mvtec.DatasetSplit.TEST,
                                              classname=category, resize=self.image_size, imagesize=self.image_size,
                                              clip_transformer=self.preprocess,
                                              divide_num=divide_num, divide_iter=divide_iter, random_seed=self.seed)
        return test_dataset

    @staticmethod
    def normalization01(img):
        """
        将图像归一化到0-1范围
        参数:
            img (numpy.ndarray): 输入图像
        返回:
            numpy.ndarray: 归一化后的图像
        """
        return (img - img.min()) / (img.max() - img.min())

    def visualization(self, image_path_list, gt_list, pr_px, category):
        """
        可视化异常检测结果，将异常图保存到指定目录
        参数:
            image_path_list (list): 图像路径列表
            gt_list (list): 真实标签列表
            pr_px (numpy.ndarray): 预测的异常图
            category (str): 类别名称
        """

        if self.vis_type == 'single_norm':
            # normalized per image
            for i, path in enumerate(image_path_list):
                anomaly_type = path.split('/')[-2]
                img_name = path.split('/')[-1]
                if anomaly_type not in ['good', 'Normal', 'ok'] and gt_list[i] != 0:
                    save_path = os.path.join(self.output_dir, category, anomaly_type)
                    os.makedirs(save_path, exist_ok=True)
                    save_path = os.path.join(save_path, img_name)

                    # 读取原始图像
                    original_img = cv2.imread(path)

                    anomaly_map = pr_px[i].squeeze()
                    anomaly_map = self.normalization01(anomaly_map) * 255
                    anomaly_map = cv2.applyColorMap(anomaly_map.astype(np.uint8), cv2.COLORMAP_JET)

                    # 调整热力图尺寸以匹配原始图像
                    h, w = original_img.shape[:2]
                    anomaly_map = cv2.resize(anomaly_map, (w, h))

                    # 叠加热力图和原图
                    overlay = cv2.addWeighted(original_img, 0.6, anomaly_map, 0.4, 0)

                    # cv2.imwrite(save_path, anomaly_map)
                    cv2.imwrite(save_path, overlay)
        else:
            # normalized all image
            pr_px = self.normalization01(pr_px)
            for i, path in enumerate(image_path_list):
                anomaly_type = path.split('/')[-2]
                img_name = path.split('/')[-1]
                save_path = os.path.join(str(self.output_dir), category, anomaly_type)
                os.makedirs(save_path, exist_ok=True)
                save_path = os.path.join(save_path, img_name)

                # 读取原始图像
                original_img = cv2.imread(path)

                anomaly_map = pr_px[i].squeeze()
                anomaly_map *= 255
                anomaly_map = cv2.applyColorMap(anomaly_map.astype(np.uint8), cv2.COLORMAP_JET)

                # 调整热力图尺寸以匹配原始图像
                h, w = original_img.shape[:2]
                anomaly_map = cv2.resize(anomaly_map, (w, h))

                # 叠加热力图和原图
                overlay = cv2.addWeighted(original_img, 0.6, anomaly_map, 0.4, 0)

                # 保存叠加后的图像
                cv2.imwrite(save_path, overlay)
                # cv2.imwrite(save_path, anomaly_map)

    def make_category_data(self, category, return_metrics=False):
        """
        处理特定类别的数据，进行异常检测并计算指标
        参数:
            category: str - 要处理的类别名称
            return_metrics: - 是否执行compute_metrics 计算
        返回:
            tuple: 包含图像级和像素级指标的元组

        """
        if category is None:
            raise ValueError('category cannot be None')
        print("Currently processing category:", category)  # 打印当前处理的类别

        # divide sub-datasets
        divide_num = self.divide_num
        anomaly_maps = torch.tensor([]).double()
        gt_list = []
        img_masks = []
        class_tokens = []
        image_path_list = []
        start_time_all = time.time()
        dataset_num = 0
        for divide_iter in range(divide_num):
            test_dataset = self.load_datasets(category, divide_num=divide_num, divide_iter=divide_iter)
            test_dataloader = torch.utils.data.DataLoader(test_dataset, batch_size=self.batch_size, shuffle=False,
                                                          num_workers=0, pin_memory=True, )
            # extract features
            patch_tokens_list = []
            subset_num = len(test_dataset)
            dataset_num += subset_num
            start_time = time.time()
            image = None
            for image_info in tqdm(test_dataloader):
                if isinstance(image_info, dict):
                    image = image_info["image"]
                    image_path_list.extend(image_info["image_path"])
                    img_masks.append(image_info["mask"])
                    gt_list.extend(list(image_info["is_anomaly"].numpy()))
                with torch.no_grad(), torch.cuda.amp.autocast():
                    input_image = image.to(torch.float).to(self.device)
                    if 'dinov2' in self.model_name:
                        patch_tokens = self.dino_model.get_intermediate_layers(x=input_image,
                                                                               n=[l - 1 for l in self.features_list],
                                                                               return_class_token=False)
                        image_features = self.dino_model(input_image)
                        patch_tokens = [patch_tokens[l].cpu() for l in range(len(self.features_list))]
                        fake_cls = [torch.zeros_like(p)[:, 0:1, :] for p in patch_tokens]
                        patch_tokens = [torch.cat([fake_cls[i], patch_tokens[i]], dim=1) for i in
                                        range(len(patch_tokens))]
                    elif 'dino' in self.model_name:
                        patch_tokens_all = self.dino_model.get_intermediate_layers(x=input_image,
                                                                                   n=max(self.features_list))
                        image_features = self.dino_model(input_image)
                        patch_tokens = [patch_tokens_all[l - 1].cpu() for l in self.features_list]
                    else:
                        # clip
                        image_features, patch_tokens = self.clip_model.encode_image(input_image, self.features_list)
                        image_features /= image_features.norm(dim=-1, keepdim=True)
                        patch_tokens = [patch_tokens[l].cpu() for l in range(len(self.features_list))]
                image_features = [image_features[bi].squeeze().cpu().numpy() for bi in range(image_features.shape[0])]
                class_tokens.extend(image_features)
                patch_tokens_list.append(patch_tokens)  # (B, L+1, C)
            end_time = time.time()
            print('extract time: {}ms per image'.format((end_time - start_time) * 1000 / subset_num))

            # LNAMD
            feature_dim = patch_tokens_list[0][0].shape[-1]
            anomaly_maps_r = torch.tensor([]).double()
            for r in self.r_list:
                start_time = time.time()
                print('aggregation degree: {}'.format(r))
                LNAMD_r = LNAMD(compute_device=self.device, r=r, feature_dim=feature_dim,
                                feature_layer=self.features_list)
                Z_layers = {}
                for im in range(len(patch_tokens_list)):
                    patch_tokens = [p.to(self.device) for p in patch_tokens_list[im]]
                    with torch.no_grad(), torch.cuda.amp.autocast():
                        features = LNAMD_r._embed(patch_tokens)
                        features /= features.norm(dim=-1, keepdim=True)
                        for l in range(len(self.features_list)):
                            # save the aggregated features
                            if str(l) not in Z_layers.keys():
                                Z_layers[str(l)] = []
                            Z_layers[str(l)].append(features[:, :, l, :])
                end_time = time.time()
                print('LNAMD-{}: {}ms per image'.format(r, (end_time - start_time) * 1000 / subset_num))

                # MSM
                anomaly_maps_l = torch.tensor([]).double()
                start_time = time.time()
                for l in Z_layers.keys():
                    # different layers
                    Z = torch.cat(Z_layers[l], dim=0).to(self.device)  # (N, L, C)
                    print('layer-{} mutual scoring...'.format(l))
                    anomaly_maps_msm = MSM(Z=Z, device=self.device, topmin_min=0, topmin_max=0.3)
                    anomaly_maps_l = torch.cat((anomaly_maps_l, anomaly_maps_msm.unsqueeze(0).cpu()), dim=0)
                anomaly_maps_l = torch.mean(anomaly_maps_l, 0)
                anomaly_maps_r = torch.cat((anomaly_maps_r, anomaly_maps_l.unsqueeze(0)), dim=0)
                end_time = time.time()
                print('MSM: {}ms per image'.format((end_time - start_time) * 1000 / subset_num))
            anomaly_maps_iter = torch.mean(anomaly_maps_r, 0).to(self.device)
            del anomaly_maps_r
            if torch.cuda.is_available():
                torch.cuda.empty_cache()

            # interpolate
            B, L = anomaly_maps_iter.shape
            H = int(np.sqrt(L))
            anomaly_maps_iter = F.interpolate(anomaly_maps_iter.view(B, 1, H, H),
                                              # size=self.image_size, mode='bicubic', align_corners=True)
                                              size=self.image_size, mode='bilinear', align_corners=True)
            anomaly_maps = torch.cat((anomaly_maps, anomaly_maps_iter.cpu()), dim=0)

        # save image features for optimizing classification
        # cls_save_path = os.path.join('./image_features/{}_{}.dat'.format(dataset, category))
        # with open(cls_save_path, 'wb') as f:
        #     pickle.dump([np.array(class_tokens)], f)
        end_time_all = time.time()
        print('MuSc: {}ms per image'.format((end_time_all - start_time_all) * 1000 / dataset_num))

        anomaly_maps = anomaly_maps.cpu().numpy()
        if torch.cuda.is_available():
            torch.cuda.empty_cache()

        B = anomaly_maps.shape[0]  # the number of unlabeled test images
        ac_score = np.array(anomaly_maps).reshape(B, -1).max(-1)
        # RsCIN
        if self.dataset == 'visa':
            k_score = [1, 8, 9]
        elif self.dataset == 'mvtec_ad':
            k_score = [1, 2, 3]
        else:
            k_score = [1, 2, 3]
        scores_cls = RsCIN(ac_score, class_tokens, k_list=k_score)

        print('computing metrics...')
        pr_sp = np.array(scores_cls)
        gt_sp = np.array(gt_list)
        gt_px = torch.cat(img_masks, dim=0).numpy().astype(np.int32)
        pr_px = np.array(anomaly_maps)
        print("Predict:", pr_sp)
        print("Ground Truth:", gt_sp)
        image_metric = None
        pixel_metric = None
        if return_metrics:
            # compute_metrics 比较耗时，后续可优化
            image_metric, pixel_metric = compute_metrics(gt_sp, pr_sp, gt_px, pr_px)
            auroc_sp, f1_sp, ap_sp = image_metric
            auroc_px, f1_px, ap_px, aupro = pixel_metric
            print('image-level, auroc:{}, f1:{}, ap:{}'.format(auroc_sp * 100, f1_sp * 100, ap_sp * 100))
            print('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(auroc_px * 100, f1_px * 100, ap_px * 100,
                                                                         aupro * 100))
        if self.vis:
            print('visualization...')
            self.visualization(image_path_list, gt_list, pr_px, category)

        # 如果 执行了compute_metrics ，则返回的  image_metric, pixel_metric 有数据，否则是None
        return image_metric, pixel_metric, image_path_list, pr_sp, gt_sp

    def save_excel_results(self, auroc_px_ls, f1_px_ls, ap_px_ls, aupro_ls, auroc_sp_ls, f1_sp_ls, ap_sp_ls,
                           auroc_px_mean, f1_px_mean, ap_px_mean, aupro_mean, auroc_sp_mean, f1_sp_mean, ap_sp_mean, ):
        if self.save_excel:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "MuSc_results"
            sheet.cell(row=1, column=2, value='auroc_px')
            sheet.cell(row=1, column=3, value='f1_px')
            sheet.cell(row=1, column=4, value='ap_px')
            sheet.cell(row=1, column=5, value='aupro')
            sheet.cell(row=1, column=6, value='auroc_sp')
            sheet.cell(row=1, column=7, value='f1_sp')
            sheet.cell(row=1, column=8, value='ap_sp')
            for col_index in range(2):
                for row_index in range(len(self.categories)):
                    if col_index == 0:
                        sheet.cell(row=row_index + 2, column=col_index + 1, value=self.categories[row_index])
                    else:
                        sheet.cell(row=row_index + 2, column=col_index + 1, value=auroc_px_ls[row_index] * 100)
                        sheet.cell(row=row_index + 2, column=col_index + 2, value=f1_px_ls[row_index] * 100)
                        sheet.cell(row=row_index + 2, column=col_index + 3, value=ap_px_ls[row_index] * 100)
                        sheet.cell(row=row_index + 2, column=col_index + 4, value=aupro_ls[row_index] * 100)
                        sheet.cell(row=row_index + 2, column=col_index + 5, value=auroc_sp_ls[row_index] * 100)
                        sheet.cell(row=row_index + 2, column=col_index + 6, value=f1_sp_ls[row_index] * 100)
                        sheet.cell(row=row_index + 2, column=col_index + 7, value=ap_sp_ls[row_index] * 100)
                    if row_index == len(self.categories) - 1:
                        if col_index == 0:
                            sheet.cell(row=row_index + 3, column=col_index + 1, value='mean')
                        else:
                            sheet.cell(row=row_index + 3, column=col_index + 1, value=auroc_px_mean * 100)
                            sheet.cell(row=row_index + 3, column=col_index + 2, value=f1_px_mean * 100)
                            sheet.cell(row=row_index + 3, column=col_index + 3, value=ap_px_mean * 100)
                            sheet.cell(row=row_index + 3, column=col_index + 4, value=aupro_mean * 100)
                            sheet.cell(row=row_index + 3, column=col_index + 5, value=auroc_sp_mean * 100)
                            sheet.cell(row=row_index + 3, column=col_index + 6, value=f1_sp_mean * 100)
                            sheet.cell(row=row_index + 3, column=col_index + 7, value=ap_sp_mean * 100)
            workbook.save(os.path.join(self.output_dir, "results.xlsx"))

    def evaluate(self):
        """
        执行完整的预测和评估
        Returns:

        """
        auroc_sp_ls = []
        f1_sp_ls = []
        ap_sp_ls = []
        auroc_px_ls = []
        f1_px_ls = []
        ap_px_ls = []
        aupro_ls = []
        # 遍历类别计算指标
        for category in self.categories:
            # 关键函数 evaluate_category_data
            image_metric, pixel_metric, _, _, _ = self.make_category_data(category=category, return_metrics=True)

            auroc_sp, f1_sp, ap_sp = image_metric
            auroc_px, f1_px, ap_px, aupro = pixel_metric
            auroc_sp_ls.append(auroc_sp)
            f1_sp_ls.append(f1_sp)
            ap_sp_ls.append(ap_sp)
            auroc_px_ls.append(auroc_px)
            f1_px_ls.append(f1_px)
            ap_px_ls.append(ap_px)
            aupro_ls.append(aupro)
        # mean
        auroc_sp_mean = sum(auroc_sp_ls) / len(auroc_sp_ls)
        f1_sp_mean = sum(f1_sp_ls) / len(f1_sp_ls)
        ap_sp_mean = sum(ap_sp_ls) / len(ap_sp_ls)
        auroc_px_mean = sum(auroc_px_ls) / len(auroc_px_ls)
        f1_px_mean = sum(f1_px_ls) / len(f1_px_ls)
        ap_px_mean = sum(ap_px_ls) / len(ap_px_ls)
        aupro_mean = sum(aupro_ls) / len(aupro_ls)

        for i, category in enumerate(self.categories):
            print('image-level, auroc:{}, f1:{}, ap:{}'.format(auroc_sp_ls[i] * 100, f1_sp_ls[i] * 100,
                                                               ap_sp_ls[i] * 100))
            print('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(auroc_px_ls[i] * 100, f1_px_ls[i] * 100,
                                                                         ap_px_ls[i] * 100, aupro_ls[i] * 100))
        print('mean')
        print('image-level, auroc:{}, f1:{}, ap:{}'.format(auroc_sp_mean * 100, f1_sp_mean * 100, ap_sp_mean * 100))
        print('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(auroc_px_mean * 100, f1_px_mean * 100,
                                                                     ap_px_mean * 100, aupro_mean * 100))

        # save in Excel
        if self.save_excel:
            self.save_excel_results(auroc_sp_ls, f1_sp_ls, ap_sp_ls, auroc_px_ls, f1_px_ls, ap_px_ls, aupro_ls,
                                    auroc_sp_mean, f1_sp_mean, ap_sp_mean, auroc_px_mean, f1_px_mean, ap_px_mean,
                                    aupro_mean)

    def predict(self):
        """
        仅执行预测
        Returns:
        """
        # 遍历类别计算指标
        for category in self.categories:
            # 关键函数 evaluate_category_data
            _, _, image_path_list, pr_sp, _ = self.make_category_data(category=category, return_metrics=False)
            for img, score in zip(image_path_list, pr_sp):
                print(img, score)
