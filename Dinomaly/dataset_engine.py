# 根据模型判断+人工复检结果，划分数据集
import random
import os
from openpyxl import load_workbook
import shutil
import numpy as np
from tqdm import tqdm
import cv2


class ImageAddAnomaly:
    def __init__(self):
        pass

    @staticmethod
    def _iamge_add_salt_and_pepper_noise(image, salt_prob=0.01, pepper_prob=0.01):
        """
        添加椒盐噪声
        Args:
            image:
            salt_prob:
            pepper_prob:

        Returns:

        """
        noisy_image = np.copy(image)
        height, width, _ = image.shape
        salt_coords = np.random.choice(width, int(salt_prob * width * height)), np.random.choice(height,
                                                                                                 int(salt_prob * width * height))
        pepper_coords = np.random.choice(width, int(pepper_prob * width * height)), np.random.choice(height,
                                                                                                     int(pepper_prob * width * height))
        noisy_image[salt_coords] = [255, 255, 255]
        noisy_image[pepper_coords] = [0, 0, 0]
        return noisy_image

    @staticmethod
    def _image_occlude(image, horizon_start_coord, vertical_start_coord):
        """
        遮挡部分图像
        Args:
            image:
            horizon_start_coord:
            vertical_start_coord:

        Returns:

        """
        occluded_image = image.copy()
        occluded_image[
            horizon_start_coord:horizon_start_coord + 100, vertical_start_coord:vertical_start_coord + 100] = [0, 0, 0]
        return occluded_image

    @staticmethod
    def _image_blur(image, ksize):
        """
        模糊图像
        Args:
            image:

        Returns:

        """
        blurred_image = cv2.GaussianBlur(image, (ksize, ksize), 0)
        return blurred_image

    @staticmethod
    def _image_add_object(image, x, y):
        """
         生成一个圆形异常目标
        Args:
            image:

        Returns:

        """
        add_circle_image = image.copy()
        anomaly = np.zeros((50, 50, 3), dtype=np.uint8)
        cv2.circle(anomaly, (25, 25), 20, (0, 0, 255), -1)
        # 将异常目标添加到图像中
        add_circle_image[y:y + anomaly.shape[0], x:x + anomaly.shape[1]] = anomaly
        return add_circle_image

    def process(self, image_file_path, save_dir):
        """
        Args:
            image_file_path:
            save_dir:
        Returns:
        """
        image = cv2.imread(image_file_path)
        image_file_name = os.path.basename(image_file_path)

        # # 加噪声
        # for i in range(5):
        #     processed_image = self._iamge_add_salt_and_pepper_noise(image, salt_prob=0.01 + random.random() * 0.01,
        #                                                             pepper_prob=0.01 + random.random() * 0.01)
        #     dest_save_file_path = os.path.join(save_dir,
        #                                        image_file_name.split('.png')[0] + f"_add_pepper_noise_{i + 1}" + ".png")
        #     cv2.imwrite(str(dest_save_file_path), processed_image)
        # 加遮挡
        for i in range(5):
            horizon_start_coord = random.randint(0, image.shape[0] - 100)
            vertical_start_coord = random.randint(0, image.shape[1] - 100)
            processed_image = self._image_occlude(image, horizon_start_coord, vertical_start_coord)
            dest_save_file_path = os.path.join(save_dir,
                                               image_file_name.split('.png')[0] + f"_occluded_{i + 1}" + ".png")
            cv2.imwrite(str(dest_save_file_path), processed_image)

        # # 加模糊
        # for i in range(5):
        #     processed_image = self._image_blur(image, ksize=5 + i * 2)
        #     dest_save_file_path = os.path.join(save_dir,
        #                                        image_file_name.split('.png')[0] + f"_blurred_{i + 1}" + ".png")
        #     cv2.imwrite(str(dest_save_file_path), processed_image)

        # 加异常图案
        for i in range(5):
            processed_image = self._image_add_object(image, x=random.randint(0, image.shape[0] - 50),
                                                     y=random.randint(0, image.shape[1] - 50))
            dest_save_file_path = os.path.join(save_dir,
                                               image_file_name.split('.png')[0] + f"_add_circle_{i + 1}" + ".png")
            cv2.imwrite(str(dest_save_file_path), processed_image)


def copy_file(source_file, destination_file):
    if os.path.exists(destination_file):
        print("File already exists!!")
        return
    try:
        # 复制文件
        shutil.copy(source_file, destination_file)
        print(f"文件 {source_file} 已成功复制到 {destination_file}")
    except FileNotFoundError:
        print(f"源文件 {source_file} 未找到。")
    except PermissionError:
        print("没有足够的权限进行文件复制操作。")
    except Exception as e:
        print(f"发生未知错误: {e}")


def sep_spec_using_pred_table(pred_table, source_folder, destination_folder):
    # 加载工作簿
    workbook = load_workbook(filename=pred_table)

    # 获取指定工作表
    sheet = workbook['Sheet1']
    # 检查目标文件夹是否存在，如果不存在则创建
    if not os.path.exists(destination_folder):
        os.makedirs(destination_folder)
    abnormal_file_dest_dir = os.path.join(destination_folder, "test", "bad")
    normal_train_file_dest_dir = os.path.join(destination_folder, "train", "good")
    normal_test_file_dest_dir = os.path.join(destination_folder, "test", "good")
    # 组合目标文件的完整路径
    if not os.path.exists(abnormal_file_dest_dir):
        os.makedirs(abnormal_file_dest_dir)
    if not os.path.exists(normal_train_file_dest_dir):
        os.makedirs(normal_train_file_dest_dir)
    if not os.path.exists(normal_test_file_dest_dir):
        os.makedirs(normal_test_file_dest_dir)

    iaa = ImageAddAnomaly()

    # 遍历工作表中的行
    for row in sheet.iter_rows(values_only=True):
        _, filename, sp_score, pred_tag = row
        if filename.endswith(".png"):
            source_file = os.path.join(source_folder, filename)
            destination_file = ""
            # 异常样本放到test/bad目录下
            if pred_tag is True:
                destination_file = os.path.join(abnormal_file_dest_dir, os.path.basename(source_file))
            # 正常样本放到train或者test/good目录下
            elif pred_tag is False:
                # 10%的概率放到test数据集下
                # 生成一个0-1的随机数
                if random.random() <= 0.1:
                    destination_file = os.path.join(normal_test_file_dest_dir, os.path.basename(source_file))
                    # 再构造一批人工异常数据
                    if os.path.exists(source_file):
                        iaa.process(source_file, abnormal_file_dest_dir)
                    else:
                        print(f'Warning: {source_file} not found, skipping augmentation')
                else:
                    destination_file = os.path.join(normal_train_file_dest_dir, os.path.basename(source_file))
            copy_file(source_file, destination_file)


def generate_ghost_ground_truth(test_dir_list):
    """
    生成幽灵 ground truth数据
    Returns:
    """
    for test_dir in test_dir_list:
        for root, dirs, files in os.walk(os.path.join(test_dir, 'test')):
            # 只处理bad文件夹
            if 'bad' not in root:
                continue
            for file in tqdm(files):
                # 如果没有ground_truth文件夹，则创建
                if not os.path.exists(os.path.join(test_dir, 'ground_truth')):
                    os.makedirs(os.path.join(test_dir, 'ground_truth'))
                # 如果没有bad文件夹，则创建
                if not os.path.exists(os.path.join(test_dir, 'ground_truth', 'bad')):
                    os.makedirs(os.path.join(test_dir, 'ground_truth', 'bad'))
                # 生成幽灵数据，复制'ghost.png' 到ground_truth文件夹
                shutil.copy('ghost.png',
                            os.path.join(test_dir, 'ground_truth', 'bad', file.split('.')[0] + '_mask.png'))


if __name__ == '__main__':
    sep_spec_using_pred_table("dinomaly_dinov3_results_qzgy.xlsx",
                              source_folder="/mnt/test/scripts/asd_for_spk/slice/qzgy/",
                              destination_folder="/mnt/test/scripts/asd_for_spk/data/spk_260123_with_manual/qzgy/")
    sep_spec_using_pred_table("dinomaly_dinov3_results_dk.xlsx",
                              source_folder="/mnt/test/scripts/asd_for_spk/slice/dk/",
                              destination_folder="/mnt/test/scripts/asd_for_spk/data/spk_260123_with_manual/dk/")
    generate_ghost_ground_truth(["/mnt/test/scripts/asd_for_spk/data/spk_260123_with_manual/qzgy/",
                                 "/mnt/test/scripts/asd_for_spk/data/spk_260123_with_manual/dk/"])
